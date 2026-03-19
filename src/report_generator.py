import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import NumberFormatDescriptor
from .config import OUTPUT_DIR, SUMMARIES_DIR
import logging

logger = logging.getLogger("REPORTER")

class ReportGenerator:
    """針對單一 PDF 生成專業報表，以及公司跨年度彙總報表"""

    def __init__(self):
        self.output_dir = OUTPUT_DIR
        self.summaries_dir = SUMMARIES_DIR

    def generate(self, company_name, summary_df, heatmap_dict, chapter_df):
        """
        為特定的公司生成獨立報表
        
        Args:
            company_name: 公司名稱
            summary_df: 包含 Root_Concept, E/S/G Count, E/S/G % 的匯總數據
            heatmap_dict: {Root_Concept: {Page_Num: Count}} 的熱點數據
            chapter_df: 由 LLM 解析出的章節範圍 DataFrame
        """
        output_path = self.output_dir / f"{company_name}_analysis.xlsx"
        
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # 1. Summary (包含 Root_Concept, E/S/G 統計)
                summary_df.to_excel(writer, index=False, sheet_name='Summary')
                
                # 設定百分比格式
                ws_summary = writer.sheets['Summary']
                self._apply_percent_format(ws_summary, summary_df)
                
                # 2. Report Structure (由 LLM 解析出的範圍)
                if not chapter_df.empty:
                    chapter_df.to_excel(writer, index=False, sheet_name='Report_Chapters')
                
                # 3. Hotspot Matrix
                if heatmap_dict:
                    # 將字典轉換為 DataFrame，Index 是 Main_Keyword, Columns 是 Page_Num
                    matrix_df = pd.DataFrame.from_dict(heatmap_dict, orient='index').fillna(0)
                    
                    # 確保索引包含所有在 summary_df 中出現的 Main_Keyword (排除總計行)，並保持順序
                    all_keywords = [k for k in summary_df['Main_Keyword'].tolist() if not k.startswith("【")]
                    matrix_df = matrix_df.reindex(all_keywords, fill_value=0)
                    
                    # 將欄位名稱 (頁碼) 轉換為整數並排序
                    matrix_df.columns = [int(c) for c in matrix_df.columns]
                    matrix_df = matrix_df.reindex(sorted(matrix_df.columns), axis=1)
                    
                    matrix_df.to_excel(writer, sheet_name='Hotspot_Matrix')
                    self._apply_heatmap(writer.sheets['Hotspot_Matrix'], matrix_df.shape[0]+1, matrix_df.shape[1]+1)
                
            display_path = str(output_path)
            if len(display_path) > 40:
                display_path = "..." + display_path[-37:]
            logger.info(f"報表已生成: {display_path}")
            return output_path
        except Exception as e:
            logger.error(f"報表生成失敗 ({company_name}): {e}")
            import traceback
            logger.error(traceback.format_exc())
            return None

    def generate_company_summary(self, company_id, all_results_df):
        """
        當同一個公司有多個年度報告時，生成一個彙總報表。
        all_results_df: 包含多個年度數據的 DataFrame，必須包含 'Year', 'Category', 'Main_Keyword' 等
        """
        output_path = self.summaries_dir / f"{company_id}_multiyear_summary.xlsx"
        
        try:
            # 僅針對特定類別總計行進行統計（例如 AI 技術、數位科技）
            summary_rows = all_results_df[all_results_df['Main_Keyword'].str.startswith("【")].copy()
            if summary_rows.empty:
                logger.warning(f"公司 {company_id} 無法生成彙總報表：找不到類別總計數據。")
                return None

            # 清理類別名稱
            summary_rows['Category_Clean'] = summary_rows['Main_Keyword'].str.replace("【", "").str.replace(" 總計】", "")
            
            # 按年份排序
            summary_rows = summary_rows.sort_values('Year')

            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # 1. 原始數據表
                summary_rows.to_excel(writer, index=False, sheet_name='All_Years_Raw')
                
                # 2. 類別趨勢表 (以 Total_Count 為主)
                # 使用 pivot_table 以防有重複年份
                pivot_total = summary_rows.pivot_table(index='Category_Clean', columns='Year', values='Total_Count', aggfunc='sum').fillna(0)
                pivot_total.to_excel(writer, sheet_name='Category_Counts_Trend')
                
                # 3. 類別佔比趨勢 (E/S/G 分佈)
                for cat in summary_rows['Category_Clean'].unique():
                    cat_data = summary_rows[summary_rows['Category_Clean'] == cat]
                    # 只取百分比欄位
                    pct_cols = ['Year', 'E_%', 'S_%', 'G_%']
                    # 確保 pct_cols 存在於欄位中
                    valid_pct_cols = [c for c in pct_cols if c in cat_data.columns]
                    cat_pct_trend = cat_data[valid_pct_cols].set_index('Year').sort_index()
                    
                    sheet_name = f'{cat[:10]}_ESG_Dist' # 限制工作表名稱長度
                    cat_pct_trend.to_excel(writer, sheet_name=sheet_name)
                    self._apply_percent_format(writer.sheets[sheet_name], cat_pct_trend)

            logger.info(f"公司年度彙總報表已生成: {output_path.name}")
            return output_path
        except Exception as e:
            logger.error(f"彙總報表生成失敗 ({company_id}): {e}")
            import traceback
            logger.error(traceback.format_exc())
            return None

    def _apply_percent_format(self, ws, df):
        """將百分比欄位設定為 Excel 的百分比格式"""
        for col_idx, col_name in enumerate(df.columns, 1):
            if '%' in col_name:
                letter = get_column_letter(col_idx)
                for row_idx in range(2, len(df) + 2):
                    ws.cell(row=row_idx, column=col_idx).number_format = '0.00%'

    def _apply_heatmap(self, ws, max_row, max_col):
        """套用 Excel 色階格式：按列(Row)進行分析，並跳過全為 0 的列"""
        color_scale_rule = ColorScaleRule(
            start_type='num', start_value=0, start_color='FFFFFF', # 0 為白色
            mid_type='percentile', mid_value=50, mid_color='FFFF00', # 50% 為黃色
            end_type='max', end_color='FF0000' # 最大值為紅色
        )
        
        for row_idx in range(2, max_row + 1):
            # 檢查該列的值，確定最大值是否大於 0
            row_max = 0
            for col_idx in range(2, max_col + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val and isinstance(val, (int, float)):
                    if val > row_max:
                        row_max = val
            
            # 如果整列最大值是 0，則不需要套用色階 (保持白色)
            if row_max == 0:
                continue
                
            first_col = get_column_letter(2)
            last_col = get_column_letter(max_col)
            addr = f"{first_col}{row_idx}:{last_col}{row_idx}"
            ws.conditional_formatting.add(addr, color_scale_rule)
